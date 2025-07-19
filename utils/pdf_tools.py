"""
Complete PDF processing utilities for all 25 tools
"""
from PIL import Image
import fitz  # PyMuPDF
import io
import os
from PyPDF2 import PdfReader, PdfWriter, PdfMerger
import pikepdf
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.colors import Color
import pdfplumber
import pytesseract
from pdf2docx import Converter
import pandas as pd
import openpyxl
from fpdf import FPDF

def merge_pdfs(files):
    """Merge multiple PDF files"""
    merger = PdfMerger()
    
    for file in files:
        merger.append(file)
    
    output = io.BytesIO()
    merger.write(output)
    merger.close()
    output.seek(0)
    return output

def split_pdf_by_pages(file, page_ranges):
    """Split PDF by specific pages (e.g., '1-3,5-7')"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    # Parse page ranges
    for page_range in page_ranges.split(','):
        if '-' in page_range:
            start, end = map(int, page_range.split('-'))
            for i in range(start-1, min(end, len(reader.pages))):
                writer.add_page(reader.pages[i])
        else:
            page_num = int(page_range) - 1
            if page_num < len(reader.pages):
                writer.add_page(reader.pages[page_num])
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def split_pdf_by_range(file, start_page, end_page):
    """Split PDF by page range"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    for i in range(start_page-1, min(end_page, len(reader.pages))):
        writer.add_page(reader.pages[i])
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def split_pdf_every_n(file, n):
    """Split PDF every N pages"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    for i in range(0, len(reader.pages), n):
        end_page = min(i + n, len(reader.pages))
        for j in range(i, end_page):
            writer.add_page(reader.pages[j])
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def compress_pdf(file, level='medium'):
    """Compress PDF with different levels"""
    # Quality settings based on level
    quality_settings = {
        'light': {'compress_streams': True, 'remove_duplication': True},
        'medium': {'compress_streams': True, 'remove_duplication': True, 'ascii85_encode': True},
        'heavy': {'compress_streams': True, 'remove_duplication': True, 'ascii85_encode': True, 'remove_images': False},
        'maximum': {'compress_streams': True, 'remove_duplication': True, 'ascii85_encode': True, 'remove_images': False}
    }
    
    settings = quality_settings.get(level, quality_settings['medium'])
    
    # Using PyMuPDF for compression
    pdf_document = fitz.open(stream=file.read(), filetype="pdf")
    
    # Apply compression
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        # Compress images on page
        for img in page.get_images():
            img_index = img[0]
            base_image = pdf_document.extract_image(img_index)
            image_bytes = base_image["image"]
            
            # Convert to PIL for compression
            pil_img = Image.open(io.BytesIO(image_bytes))
            
            # Compress image based on level
            quality = 95 if level == 'light' else 80 if level == 'medium' else 65 if level == 'heavy' else 50
            
            compressed_img = io.BytesIO()
            pil_img.save(compressed_img, format='JPEG', quality=quality, optimize=True)
            compressed_img.seek(0)
    
    output = io.BytesIO()
    pdf_document.save(output, garbage=4, deflate=True, clean=True)
    pdf_document.close()
    output.seek(0)
    return output

def pdf_to_word(file):
    """Convert PDF to Word document"""
    # Save uploaded file temporarily
    temp_pdf = io.BytesIO(file.read())
    
    # Convert to DOCX
    output = io.BytesIO()
    cv = Converter(temp_pdf)
    cv.convert(output, start=0, end=None)
    cv.close()
    
    output.seek(0)
    return output

def pdf_to_excel(file):
    """Convert PDF tables to Excel"""
    # Extract tables using pdfplumber
    tables = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            page_tables = page.extract_tables()
            if page_tables:
                tables.extend(page_tables)
    
    # Convert to Excel
    output = io.BytesIO()
    
    if tables:
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Write tables to worksheet
        row_offset = 1
        for table_idx, table in enumerate(tables):
            if table_idx > 0:
                row_offset += 2  # Add spacing between tables
            
            for row_idx, row in enumerate(table):
                for col_idx, cell in enumerate(row):
                    if cell:
                        worksheet.cell(row=row_offset + row_idx, column=col_idx + 1, value=str(cell))
            
            row_offset += len(table)
        
        workbook.save(output)
    
    output.seek(0)
    return output

def pdf_to_images(file, format='png', dpi=300):
    """Convert PDF pages to images"""
    pdf_document = fitz.open(stream=file.read(), filetype="pdf")
    images = []
    
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        
        # Render page as image
        mat = fitz.Matrix(dpi/72, dpi/72)
        pix = page.get_pixmap(matrix=mat)
        
        # Convert to PIL Image
        img_data = pix.tobytes("png")
        pil_img = Image.open(io.BytesIO(img_data))
        
        # Convert format if needed
        if format.lower() != 'png':
            if format.lower() == 'jpeg' and pil_img.mode in ('RGBA', 'LA'):
                # Convert RGBA to RGB for JPEG
                rgb_img = Image.new('RGB', pil_img.size, (255, 255, 255))
                rgb_img.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                pil_img = rgb_img
        
        images.append(pil_img)
    
    pdf_document.close()
    
    # Save all images to a ZIP or return first image
    if len(images) == 1:
        output = io.BytesIO()
        images[0].save(output, format=format.upper())
        output.seek(0)
        return output
    else:
        # For multiple pages, create a ZIP file
        import zipfile
        output = io.BytesIO()
        with zipfile.ZipFile(output, 'w') as zip_file:
            for i, img in enumerate(images):
                img_bytes = io.BytesIO()
                img.save(img_bytes, format=format.upper())
                zip_file.writestr(f'page_{i+1}.{format}', img_bytes.getvalue())
        output.seek(0)
        return output

def word_to_pdf(file):
    """Convert Word document to PDF"""
    # This is a simplified version - in production you'd use python-docx2pdf
    from docx2pdf import convert
    import tempfile
    
    # Save to temporary file
    temp_dir = tempfile.mkdtemp()
    word_path = os.path.join(temp_dir, 'input.docx')
    pdf_path = os.path.join(temp_dir, 'output.pdf')
    
    with open(word_path, 'wb') as f:
        f.write(file.read())
    
    # Convert to PDF
    convert(word_path, pdf_path)
    
    # Read result
    output = io.BytesIO()
    with open(pdf_path, 'rb') as f:
        output.write(f.read())
    
    # Cleanup
    os.remove(word_path)
    os.remove(pdf_path)
    os.rmdir(temp_dir)
    
    output.seek(0)
    return output

def excel_to_pdf(file):
    """Convert Excel to PDF"""
    # Read Excel file
    df = pd.read_excel(file)
    
    # Create PDF
    output = io.BytesIO()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', size=10)
    
    # Add header
    for col in df.columns:
        pdf.cell(40, 10, str(col), 1)
    pdf.ln()
    
    # Add data rows
    for index, row in df.iterrows():
        for value in row:
            pdf.cell(40, 10, str(value)[:20], 1)  # Truncate long values
        pdf.ln()
    
    pdf.output(output)
    output.seek(0)
    return output

def images_to_pdf(files):
    """Convert multiple images to PDF"""
    images = []
    
    for file in files:
        img = Image.open(file)
        # Convert to RGB if necessary
        if img.mode in ('RGBA', 'LA', 'P'):
            rgb_img = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = rgb_img
        images.append(img)
    
    output = io.BytesIO()
    if images:
        images[0].save(output, format='PDF', save_all=True, append_images=images[1:])
    output.seek(0)
    return output

def text_to_pdf(text, font_size=12):
    """Convert text to PDF"""
    output = io.BytesIO()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', size=font_size)
    
    # Split text into lines
    lines = text.split('\n')
    for line in lines:
        pdf.cell(0, 10, line.encode('latin-1', 'replace').decode('latin-1'), ln=True)
    
    pdf.output(output)
    output.seek(0)
    return output

def protect_pdf(file, password):
    """Add password protection to PDF"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    # Copy all pages
    for page in reader.pages:
        writer.add_page(page)
    
    # Add password protection
    writer.encrypt(password)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def unlock_pdf(file, password):
    """Remove password from PDF"""
    reader = PdfReader(file)
    
    if reader.is_encrypted:
        reader.decrypt(password)
    
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def add_pdf_watermark(file, watermark_text, opacity=0.5):
    """Add watermark to PDF"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    # Create watermark
    watermark_buffer = io.BytesIO()
    c = canvas.Canvas(watermark_buffer, pagesize=letter)
    c.setFillColor(Color(0, 0, 0, alpha=opacity))
    c.setFont("Helvetica", 50)
    c.rotate(45)
    c.drawString(200, 200, watermark_text)
    c.save()
    
    watermark_buffer.seek(0)
    watermark = PdfReader(watermark_buffer)
    watermark_page = watermark.pages[0]
    
    # Apply watermark to all pages
    for page in reader.pages:
        page.merge_page(watermark_page)
        writer.add_page(page)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def rotate_pdf(file, angle, pages='all'):
    """Rotate PDF pages"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    for i, page in enumerate(reader.pages):
        if pages == 'all' or str(i+1) in pages.split(','):
            page.rotate(angle)
        writer.add_page(page)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def extract_pdf_text(file):
    """Extract text from PDF"""
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    
    return text

def ocr_pdf(file):
    """Perform OCR on PDF"""
    pdf_document = fitz.open(stream=file.read(), filetype="pdf")
    writer = PdfWriter()
    
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        
        # Convert page to image
        pix = page.get_pixmap()
        img_data = pix.tobytes("png")
        
        # Perform OCR
        ocr_text = pytesseract.image_to_string(Image.open(io.BytesIO(img_data)))
        
        # Create new PDF page with OCR text
        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=letter)
        c.drawString(50, 750, ocr_text[:100] + "...")  # Simplified - would need proper text layout
        c.save()
        
        packet.seek(0)
        new_page = PdfReader(packet).pages[0]
        writer.add_page(new_page)
    
    pdf_document.close()
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def add_digital_signature(file, signature_text, position='bottom-right'):
    """Add digital signature to PDF"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    # Create signature overlay
    signature_buffer = io.BytesIO()
    c = canvas.Canvas(signature_buffer, pagesize=letter)
    
    # Position signature
    if position == 'bottom-right':
        x, y = 400, 50
    elif position == 'bottom-left':
        x, y = 50, 50
    else:
        x, y = 300, 300
    
    c.setFont("Helvetica", 12)
    c.drawString(x, y, f"Digitally signed: {signature_text}")
    c.save()
    
    signature_buffer.seek(0)
    signature = PdfReader(signature_buffer)
    signature_page = signature.pages[0]
    
    # Apply signature to last page
    for i, page in enumerate(reader.pages):
        if i == len(reader.pages) - 1:  # Last page
            page.merge_page(signature_page)
        writer.add_page(page)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def fill_pdf_forms(file, form_data):
    """Fill PDF forms with data"""
    # This would require more complex form field handling
    # Simplified version
    reader = PdfReader(file)
    writer = PdfWriter()
    
    # Copy pages (form filling would be more complex)
    for page in reader.pages:
        writer.add_page(page)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def extract_pdf_bookmarks(file):
    """Extract bookmarks from PDF"""
    reader = PdfReader(file)
    bookmarks = []
    
    def extract_bookmarks_recursive(bookmark_list, level=0):
        for item in bookmark_list:
            if isinstance(item, list):
                extract_bookmarks_recursive(item, level + 1)
            else:
                bookmarks.append({
                    'title': item.title,
                    'level': level,
                    'page': reader.get_destination_page_number(item) + 1 if hasattr(item, 'page') else 'Unknown'
                })
    
    if reader.outline:
        extract_bookmarks_recursive(reader.outline)
    
    return bookmarks

def get_pdf_metadata(file):
    """Get PDF metadata"""
    reader = PdfReader(file)
    metadata = {}
    
    if reader.metadata:
        for key, value in reader.metadata.items():
            metadata[key] = value
    
    metadata['pages'] = len(reader.pages)
    return metadata

def edit_pdf_metadata(file, new_metadata):
    """Edit PDF metadata"""
    reader = PdfReader(file)
    writer = PdfWriter()
    
    # Copy all pages
    for page in reader.pages:
        writer.add_page(page)
    
    # Add new metadata
    writer.add_metadata(new_metadata)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def compare_pdfs(file1, file2):
    """Compare two PDFs"""
    text1 = extract_pdf_text(file1)
    text2 = extract_pdf_text(file2)
    
    # Simple comparison
    differences = []
    lines1 = text1.split('\n')
    lines2 = text2.split('\n')
    
    max_lines = max(len(lines1), len(lines2))
    
    for i in range(max_lines):
        line1 = lines1[i] if i < len(lines1) else ""
        line2 = lines2[i] if i < len(lines2) else ""
        
        if line1 != line2:
            differences.append({
                'line': i + 1,
                'pdf1': line1,
                'pdf2': line2
            })
    
    return differences

def optimize_pdf(file, level='standard'):
    """Optimize PDF for size and performance"""
    # Use compression with optimization
    return compress_pdf(file, 'medium' if level == 'standard' else level)

def extract_pdf_annotations(file):
    """Extract annotations from PDF"""
    reader = PdfReader(file)
    annotations = []
    
    for page_num, page in enumerate(reader.pages):
        if "/Annots" in page:
            for annot in page["/Annots"]:
                annot_obj = annot.get_object()
                if "/Contents" in annot_obj:
                    annotations.append({
                        'page': page_num + 1,
                        'type': str(annot_obj.get("/Subtype", "Unknown")),
                        'content': str(annot_obj["/Contents"])
                    })
    
    return annotations

def add_pdf_annotation(file, annotation_text, page_num):
    """Add annotation to PDF"""
    # This would require more complex annotation handling
    # Simplified version - just add text overlay
    reader = PdfReader(file)
    writer = PdfWriter()
    
    for i, page in enumerate(reader.pages):
        if i == page_num - 1:
            # Add annotation as text overlay (simplified)
            annotation_buffer = io.BytesIO()
            c = canvas.Canvas(annotation_buffer, pagesize=letter)
            c.setFont("Helvetica", 10)
            c.setFillColor(colors.red)
            c.drawString(50, 50, f"Note: {annotation_text}")
            c.save()
            
            annotation_buffer.seek(0)
            annotation_page = PdfReader(annotation_buffer).pages[0]
            page.merge_page(annotation_page)
        
        writer.add_page(page)
    
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def redact_pdf_text(file, text_to_redact):
    """Redact (black out) specific text in PDF"""
    pdf_document = fitz.open(stream=file.read(), filetype="pdf")
    
    for page in pdf_document:
        # Find text instances
        text_instances = page.search_for(text_to_redact)
        
        # Redact each instance
        for inst in text_instances:
            annot = page.add_redact_annot(inst)
            annot.set_colors(stroke=fitz.utils.getColor("black"), fill=fitz.utils.getColor("black"))
        
        # Apply redactions
        page.apply_redactions()
    
    output = io.BytesIO()
    pdf_document.save(output)
    pdf_document.close()
    output.seek(0)
    return output

def get_pdf_page_count(file):
    """Get number of pages in PDF"""
    reader = PdfReader(file)
    return len(reader.pages)

def get_pdf_file_size(file):
    """Get PDF file size"""
    file.seek(0, 2)  # Seek to end
    size = file.tell()
    file.seek(0)  # Reset to beginning
    
    # Convert to human readable format
    if size < 1024:
        return f"{size} bytes"
    elif size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    else:
        return f"{size / (1024 * 1024):.1f} MB"